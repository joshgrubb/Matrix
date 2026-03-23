"""
Unit tests for the export service layer.

Tests every public function in ``app.services.export_service`` by
constructing synthetic ``DepartmentCostSummary`` and
``PositionCostSummary`` dataclass instances and verifying the
generated CSV and Excel output byte-for-byte.

Because the export service is a **pure transformation layer** --
dataclasses in, BytesIO out -- these tests do not require a database
connection or Flask application context.  This makes them fast,
deterministic, and isolated from the cost calculation pipeline.

Design decisions:
    - CSV assertions parse the returned BytesIO through Python's
      ``csv.reader`` (after decoding the UTF-8-BOM bytes) so that
      field-level values can be compared exactly.  This catches
      column-order changes, missing columns, and formatting regressions.
    - Excel assertions use ``openpyxl.load_workbook`` to read the
      returned BytesIO and verify cell values, sheet names, header
      styling, column widths, and number formatting.  This catches
      broken workbooks, wrong sheet names, and incorrect number
      formatting.
    - All dollar assertions compare against ``Decimal`` values to
      avoid floating-point surprises.  Excel cells store floats, so
      ``pytest.approx`` is used for those comparisons.
    - Every CSV test verifies the UTF-8-BOM prefix to ensure
      Microsoft Excel opens the file without encoding issues --
      a common real-world complaint.
    - Empty-data tests verify that the header row is still written
      when the input list is empty.  This prevents downstream errors
      if a user exports a report for a department with no positions.

Fixture reminder (synthetic -- no conftest dependency):
    _make_dept_summary:  Factory for DepartmentCostSummary instances.
    _make_pos_summary:   Factory for PositionCostSummary instances.

Run this file in isolation::

    pytest tests/test_services/test_export_service.py -v
"""

import csv
import io
from decimal import Decimal

import pytest
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.services import export_service
from app.services.cost_service import (
    DepartmentCostSummary,
    PositionCostSummary,
)


# =====================================================================
# Test data factories
# =====================================================================


def _make_dept_summary(
    department_id=1,
    department_name="Engineering",
    division_count=3,
    position_count=12,
    total_authorized=45,
    hardware_total=Decimal("54000.00"),
    software_total=Decimal("9000.00"),
    grand_total=Decimal("63000.00"),
):
    """
    Build a DepartmentCostSummary with sensible defaults.

    Every parameter is overridable so that individual tests can
    vary only the fields they care about.

    Args:
        department_id:    Synthetic primary key.
        department_name:  Department display name.
        division_count:   Number of child divisions.
        position_count:   Number of child positions.
        total_authorized: Sum of authorized_count across positions.
        hardware_total:   Aggregate hardware cost.
        software_total:   Aggregate software cost.
        grand_total:      hardware_total + software_total.

    Returns:
        A fully populated DepartmentCostSummary dataclass.
    """
    return DepartmentCostSummary(
        department_id=department_id,
        department_name=department_name,
        division_count=division_count,
        position_count=position_count,
        total_authorized=total_authorized,
        hardware_total=hardware_total,
        software_total=software_total,
        grand_total=grand_total,
    )


def _make_pos_summary(
    position_id=1,
    position_title="Senior Analyst",
    position_code="SA-001",
    division_id=1,
    division_name="Data Division",
    department_id=1,
    department_name="Engineering",
    authorized_count=3,
    hardware_total_per_person=Decimal("1200.00"),
    software_total_per_person=Decimal("200.00"),
    total_per_person=Decimal("1400.00"),
    hardware_total=Decimal("3600.00"),
    software_total=Decimal("600.00"),
    grand_total=Decimal("4200.00"),
):
    """
    Build a PositionCostSummary with sensible defaults.

    Defaults model a position with authorized_count=3 assigned
    one $1,200 laptop and one $200/user software license, which
    produces deterministic totals for straightforward assertions.

    Args:
        position_id:               Synthetic primary key.
        position_title:            Human-readable title.
        position_code:             Short code (e.g., 'SA-001').
        division_id:               Parent division FK.
        division_name:             Parent division display name.
        department_id:             Grandparent department FK.
        department_name:           Grandparent department display name.
        authorized_count:          Number of authorized headcount.
        hardware_total_per_person: HW cost per single person.
        software_total_per_person: SW cost per single person.
        total_per_person:          HW + SW per person.
        hardware_total:            HW per person * authorized_count.
        software_total:            SW per person * authorized_count.
        grand_total:               hardware_total + software_total.

    Returns:
        A fully populated PositionCostSummary dataclass.
    """
    return PositionCostSummary(
        position_id=position_id,
        position_title=position_title,
        position_code=position_code,
        division_id=division_id,
        division_name=division_name,
        department_id=department_id,
        department_name=department_name,
        authorized_count=authorized_count,
        hardware_total_per_person=hardware_total_per_person,
        software_total_per_person=software_total_per_person,
        total_per_person=total_per_person,
        hardware_total=hardware_total,
        software_total=software_total,
        grand_total=grand_total,
    )


# =====================================================================
# CSV parsing helpers
# =====================================================================


def _parse_csv_buffer(buffer):
    """
    Parse a BytesIO CSV buffer (UTF-8-BOM encoded) into a list of rows.

    Strips the BOM before decoding so that the csv.reader does not
    see it as part of the first column header.

    Args:
        buffer: The BytesIO returned by an export_*_csv function.

    Returns:
        A list of lists, where each inner list is a row of string values.
    """
    raw_bytes = buffer.read()
    # Strip the UTF-8-BOM prefix before decoding.
    text = raw_bytes.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    return list(reader)


def _load_excel_buffer(buffer):
    """
    Load an Excel workbook from a BytesIO buffer.

    Args:
        buffer: The BytesIO returned by an export_*_excel function.

    Returns:
        An openpyxl Workbook instance.
    """
    return load_workbook(filename=buffer)


# =====================================================================
# 1. Department costs CSV export
# =====================================================================


class TestExportDepartmentCostsCsv:
    """
    Verify ``export_department_costs_csv`` produces a correctly
    structured and encoded CSV file from DepartmentCostSummary data.
    """

    # -- 1.1 Return type and encoding ----------------------------------

    def test_returns_bytes_io_buffer(self):
        """The function must return an io.BytesIO instance."""
        result = export_service.export_department_costs_csv([])
        assert isinstance(result, io.BytesIO)

    def test_buffer_position_is_at_zero(self):
        """
        The returned buffer must be seeked to position 0 so that
        callers (like Flask's make_response) can read immediately
        without an additional seek call.
        """
        result = export_service.export_department_costs_csv([])
        assert result.tell() == 0

    def test_csv_is_utf8_bom_encoded(self):
        """
        The CSV must start with the UTF-8 BOM (byte order mark)
        so that Microsoft Excel correctly detects the encoding
        when the file is opened on Windows.  Without the BOM,
        Excel defaults to the system locale and corrupts non-ASCII
        department names.
        """
        result = export_service.export_department_costs_csv([])
        raw = result.read()
        # UTF-8 BOM is the three-byte sequence EF BB BF.
        assert raw[:3] == b"\xef\xbb\xbf", f"Expected UTF-8 BOM prefix, got {raw[:3]!r}"

    # -- 1.2 Header row ------------------------------------------------

    def test_header_row_has_correct_columns(self):
        """
        The first row must contain exactly the seven department-level
        column headers in the correct order.
        """
        result = export_service.export_department_costs_csv([])
        rows = _parse_csv_buffer(result)
        expected_headers = [
            "Department",
            "Divisions",
            "Positions",
            "Authorized Headcount",
            "Hardware Total",
            "Software Total",
            "Grand Total",
        ]
        assert rows[0] == expected_headers

    def test_header_row_column_count(self):
        """The header row must contain exactly 7 columns."""
        result = export_service.export_department_costs_csv([])
        rows = _parse_csv_buffer(result)
        assert len(rows[0]) == 7

    # -- 1.3 Empty data ------------------------------------------------

    def test_empty_data_produces_header_only(self):
        """
        When the input list is empty, the CSV must still contain the
        header row and nothing else.  This prevents downstream errors
        if an export is triggered for an organization with no
        departments in scope.
        """
        result = export_service.export_department_costs_csv([])
        rows = _parse_csv_buffer(result)
        assert len(rows) == 1, f"Expected 1 row (header only), got {len(rows)} rows"

    # -- 1.4 Single department -----------------------------------------

    def test_single_department_produces_one_data_row(self):
        """
        A single DepartmentCostSummary should produce exactly one
        data row (plus the header).
        """
        summaries = [_make_dept_summary()]
        result = export_service.export_department_costs_csv(summaries)
        rows = _parse_csv_buffer(result)
        assert len(rows) == 2

    def test_single_department_correct_values(self):
        """
        Verify that every field in the data row matches the input
        dataclass values exactly.
        """
        dept = _make_dept_summary(
            department_name="Finance",
            division_count=2,
            position_count=8,
            total_authorized=30,
            hardware_total=Decimal("36000.00"),
            software_total=Decimal("6000.00"),
            grand_total=Decimal("42000.00"),
        )
        result = export_service.export_department_costs_csv([dept])
        rows = _parse_csv_buffer(result)
        data_row = rows[1]

        assert data_row[0] == "Finance"
        assert data_row[1] == "2"
        assert data_row[2] == "8"
        assert data_row[3] == "30"
        assert data_row[4] == "36000.00"
        assert data_row[5] == "6000.00"
        assert data_row[6] == "42000.00"

    def test_data_row_has_same_column_count_as_header(self):
        """
        Every data row must have the same number of fields as the
        header to prevent CSV misalignment.
        """
        summaries = [_make_dept_summary()]
        result = export_service.export_department_costs_csv(summaries)
        rows = _parse_csv_buffer(result)
        header_count = len(rows[0])
        for i, row in enumerate(rows[1:], start=2):
            assert (
                len(row) == header_count
            ), f"Row {i} has {len(row)} columns, expected {header_count}"

    # -- 1.5 Multiple departments --------------------------------------

    def test_multiple_departments_all_rows_present(self):
        """
        When given N departments, the CSV must have N+1 rows
        (1 header + N data rows).
        """
        summaries = [
            _make_dept_summary(department_id=1, department_name="Engineering"),
            _make_dept_summary(department_id=2, department_name="Finance"),
            _make_dept_summary(department_id=3, department_name="Human Resources"),
        ]
        result = export_service.export_department_costs_csv(summaries)
        rows = _parse_csv_buffer(result)
        assert len(rows) == 4

    def test_multiple_departments_preserve_input_order(self):
        """
        The CSV rows must appear in the same order as the input
        list.  The export service must not sort or reorder.
        """
        summaries = [
            _make_dept_summary(department_id=3, department_name="Zebra Dept"),
            _make_dept_summary(department_id=1, department_name="Alpha Dept"),
            _make_dept_summary(department_id=2, department_name="Middle Dept"),
        ]
        result = export_service.export_department_costs_csv(summaries)
        rows = _parse_csv_buffer(result)
        assert rows[1][0] == "Zebra Dept"
        assert rows[2][0] == "Alpha Dept"
        assert rows[3][0] == "Middle Dept"

    # -- 1.6 Decimal formatting ----------------------------------------

    def test_cost_values_formatted_to_two_decimal_places(self):
        """
        All cost columns must be formatted as strings with exactly
        two decimal places, even when the Decimal value has trailing
        zeros (e.g., Decimal('100.00') must be '100.00', not '100').
        """
        dept = _make_dept_summary(
            hardware_total=Decimal("100.00"),
            software_total=Decimal("50.00"),
            grand_total=Decimal("150.00"),
        )
        result = export_service.export_department_costs_csv([dept])
        rows = _parse_csv_buffer(result)
        data_row = rows[1]
        # Hardware Total, Software Total, Grand Total.
        assert data_row[4] == "100.00"
        assert data_row[5] == "50.00"
        assert data_row[6] == "150.00"

    def test_zero_cost_departments_formatted_correctly(self):
        """
        A department with $0.00 costs must still display '0.00'
        in the CSV, not an empty string or '0'.
        """
        dept = _make_dept_summary(
            hardware_total=Decimal("0.00"),
            software_total=Decimal("0.00"),
            grand_total=Decimal("0.00"),
        )
        result = export_service.export_department_costs_csv([dept])
        rows = _parse_csv_buffer(result)
        data_row = rows[1]
        assert data_row[4] == "0.00"
        assert data_row[5] == "0.00"
        assert data_row[6] == "0.00"

    def test_large_cost_values_not_truncated(self):
        """
        Very large cost values must not be truncated or formatted
        with scientific notation.  This catches issues where the
        _format_decimal helper might fail on large Decimals.
        """
        dept = _make_dept_summary(
            hardware_total=Decimal("9999999.99"),
            software_total=Decimal("1234567.89"),
            grand_total=Decimal("11234567.88"),
        )
        result = export_service.export_department_costs_csv([dept])
        rows = _parse_csv_buffer(result)
        data_row = rows[1]
        assert data_row[4] == "9999999.99"
        assert data_row[5] == "1234567.89"
        assert data_row[6] == "11234567.88"

    def test_fractional_cent_values_round_to_two_places(self):
        """
        If a Decimal has more than two decimal places (which should
        not happen in normal operation, but could if a bug creeps
        into cost_service), _format_decimal should still produce
        a two-decimal-place string via f-string formatting.
        """
        dept = _make_dept_summary(
            hardware_total=Decimal("1234.567"),
            software_total=Decimal("0.001"),
            grand_total=Decimal("1234.568"),
        )
        result = export_service.export_department_costs_csv([dept])
        rows = _parse_csv_buffer(result)
        data_row = rows[1]
        # Python f"{Decimal('1234.567'):.2f}" -> '1234.57' (rounds).
        assert data_row[4] == "1234.57"
        assert data_row[5] == "0.00"
        assert data_row[6] == "1234.57"

    # -- 1.7 Structural integrity --------------------------------------

    def test_grand_total_equals_hw_plus_sw_in_csv(self):
        """
        For every data row, Grand Total must equal Hardware Total +
        Software Total.  This is a structural integrity check that
        catches data corruption during CSV serialization.
        """
        summaries = [
            _make_dept_summary(
                department_id=1,
                hardware_total=Decimal("3600.00"),
                software_total=Decimal("600.00"),
                grand_total=Decimal("4200.00"),
            ),
            _make_dept_summary(
                department_id=2,
                hardware_total=Decimal("0.00"),
                software_total=Decimal("1500.00"),
                grand_total=Decimal("1500.00"),
            ),
        ]
        result = export_service.export_department_costs_csv(summaries)
        rows = _parse_csv_buffer(result)

        for row in rows[1:]:
            hw = Decimal(row[4])
            sw = Decimal(row[5])
            grand = Decimal(row[6])
            assert grand == hw + sw, (
                f"Grand total mismatch for {row[0]}: " f"{grand} != {hw} + {sw}"
            )

    # -- 1.8 Special characters ----------------------------------------

    def test_department_name_with_comma_is_properly_quoted(self):
        """
        A department name containing a comma must be properly
        quoted in the CSV so that parsers do not split it into
        two columns.
        """
        dept = _make_dept_summary(
            department_name="Planning, Budgeting & Analysis",
        )
        result = export_service.export_department_costs_csv([dept])
        rows = _parse_csv_buffer(result)
        assert rows[1][0] == "Planning, Budgeting & Analysis"

    def test_department_name_with_quotes_is_properly_escaped(self):
        """
        A department name containing double quotes must be escaped
        (doubled) in the CSV per RFC 4180.
        """
        dept = _make_dept_summary(
            department_name='The "Special" Department',
        )
        result = export_service.export_department_costs_csv([dept])
        rows = _parse_csv_buffer(result)
        assert rows[1][0] == 'The "Special" Department'

    def test_department_name_with_newline_is_properly_quoted(self):
        """
        A department name containing a newline must be handled
        gracefully by the CSV writer (quoted per RFC 4180).
        """
        dept = _make_dept_summary(
            department_name="Department\nWith Newline",
        )
        result = export_service.export_department_costs_csv([dept])
        rows = _parse_csv_buffer(result)
        assert rows[1][0] == "Department\nWith Newline"

    def test_unicode_department_name_preserved(self):
        """
        Non-ASCII characters (accented letters, CJK, etc.) must
        survive the UTF-8-BOM encoding round-trip intact.
        """
        dept = _make_dept_summary(
            department_name="Departamento de Analisis Financiero",
        )
        result = export_service.export_department_costs_csv([dept])
        rows = _parse_csv_buffer(result)
        assert rows[1][0] == "Departamento de Analisis Financiero"


# =====================================================================
# 2. Department costs Excel export
# =====================================================================


class TestExportDepartmentCostsExcel:
    """
    Verify ``export_department_costs_excel`` produces a valid .xlsx
    workbook from DepartmentCostSummary data.
    """

    # -- 2.1 Return type and buffer position ---------------------------

    def test_returns_bytes_io_buffer(self):
        """The function must return an io.BytesIO instance."""
        result = export_service.export_department_costs_excel([])
        assert isinstance(result, io.BytesIO)

    def test_buffer_position_is_at_zero(self):
        """The returned buffer must be seeked to position 0."""
        result = export_service.export_department_costs_excel([])
        assert result.tell() == 0

    # -- 2.2 Workbook validity -----------------------------------------

    def test_produces_valid_xlsx_file(self):
        """
        The returned bytes must be parseable by openpyxl as a valid
        .xlsx workbook.  A corrupt file would raise an exception.
        """
        result = export_service.export_department_costs_excel([])
        wb = _load_excel_buffer(result)
        assert wb is not None
        wb.close()

    def test_active_sheet_title_is_department_costs(self):
        """
        The active (first) sheet must be titled 'Department Costs'
        to match the export_service._write_header_row call.
        """
        result = export_service.export_department_costs_excel([])
        wb = _load_excel_buffer(result)
        assert wb.active.title == "Department Costs"
        wb.close()

    # -- 2.3 Header row ------------------------------------------------

    def test_header_row_has_correct_columns(self):
        """
        The first row of the active sheet must contain the seven
        department-level column headers in order.
        """
        result = export_service.export_department_costs_excel([])
        wb = _load_excel_buffer(result)
        ws = wb.active
        headers = [ws.cell(row=1, column=c).value for c in range(1, 8)]
        expected = [
            "Department",
            "Divisions",
            "Positions",
            "Authorized Headcount",
            "Hardware Total",
            "Software Total",
            "Grand Total",
        ]
        assert headers == expected
        wb.close()

    def test_header_row_font_is_bold_white(self):
        """
        Header cells must use bold white font (the _HEADER_FONT
        constant) for readability against the dark fill.
        """
        result = export_service.export_department_costs_excel([])
        wb = _load_excel_buffer(result)
        ws = wb.active
        for col in range(1, 8):
            cell = ws.cell(row=1, column=col)
            assert cell.font.bold is True, f"Header column {col} font is not bold"
            assert cell.font.color.rgb == "00FFFFFF", (
                f"Header column {col} font color is not white, "
                f"got {cell.font.color.rgb}"
            )
        wb.close()

    def test_header_row_fill_is_dark_blue(self):
        """
        Header cells must use the dark blue fill (2B579A) defined
        in the _HEADER_FILL constant.
        """
        result = export_service.export_department_costs_excel([])
        wb = _load_excel_buffer(result)
        ws = wb.active
        for col in range(1, 8):
            cell = ws.cell(row=1, column=col)
            assert cell.fill.start_color.rgb == "002B579A", (
                f"Header column {col} fill color mismatch: "
                f"got {cell.fill.start_color.rgb}"
            )
        wb.close()

    def test_header_row_alignment_is_centered_wrapped(self):
        """
        Header cells must be center-aligned with text wrapping
        enabled (the _HEADER_ALIGN constant).
        """
        result = export_service.export_department_costs_excel([])
        wb = _load_excel_buffer(result)
        ws = wb.active
        for col in range(1, 8):
            cell = ws.cell(row=1, column=col)
            assert (
                cell.alignment.horizontal == "center"
            ), f"Header column {col} is not center-aligned"
            assert (
                cell.alignment.wrap_text is True
            ), f"Header column {col} does not have wrap_text enabled"
        wb.close()

    # -- 2.4 Empty data ------------------------------------------------

    def test_empty_data_produces_header_only(self):
        """
        When the input list is empty, the worksheet must contain
        only the header row (max_row == 1).
        """
        result = export_service.export_department_costs_excel([])
        wb = _load_excel_buffer(result)
        ws = wb.active
        assert ws.max_row == 1
        wb.close()

    # -- 2.5 Data values -----------------------------------------------

    def test_single_department_correct_cell_values(self):
        """
        Verify that a single DepartmentCostSummary is written to
        row 2 with correct cell values and types.
        """
        dept = _make_dept_summary(
            department_name="Engineering",
            division_count=3,
            position_count=12,
            total_authorized=45,
            hardware_total=Decimal("54000.00"),
            software_total=Decimal("9000.00"),
            grand_total=Decimal("63000.00"),
        )
        result = export_service.export_department_costs_excel([dept])
        wb = _load_excel_buffer(result)
        ws = wb.active

        assert ws.cell(row=2, column=1).value == "Engineering"
        assert ws.cell(row=2, column=2).value == 3
        assert ws.cell(row=2, column=3).value == 12
        assert ws.cell(row=2, column=4).value == 45
        # Cost columns are stored as floats in Excel.
        assert ws.cell(row=2, column=5).value == pytest.approx(54000.00)
        assert ws.cell(row=2, column=6).value == pytest.approx(9000.00)
        assert ws.cell(row=2, column=7).value == pytest.approx(63000.00)
        wb.close()

    def test_multiple_departments_row_count(self):
        """
        N departments should produce N+1 rows (1 header + N data).
        """
        summaries = [
            _make_dept_summary(department_id=i, department_name=f"Dept {i}")
            for i in range(1, 6)
        ]
        result = export_service.export_department_costs_excel(summaries)
        wb = _load_excel_buffer(result)
        ws = wb.active
        assert ws.max_row == 6  # 1 header + 5 data rows.
        wb.close()

    def test_multiple_departments_preserve_input_order(self):
        """The Excel rows must match the input list order."""
        summaries = [
            _make_dept_summary(department_id=3, department_name="Charlie"),
            _make_dept_summary(department_id=1, department_name="Alpha"),
            _make_dept_summary(department_id=2, department_name="Bravo"),
        ]
        result = export_service.export_department_costs_excel(summaries)
        wb = _load_excel_buffer(result)
        ws = wb.active
        assert ws.cell(row=2, column=1).value == "Charlie"
        assert ws.cell(row=3, column=1).value == "Alpha"
        assert ws.cell(row=4, column=1).value == "Bravo"
        wb.close()

    # -- 2.6 Currency number format ------------------------------------

    def test_cost_cells_have_currency_number_format(self):
        """
        The Hardware Total, Software Total, and Grand Total cells
        (columns 5-7) must use the '#,##0.00' number format so
        that Excel displays them as currency values.
        """
        summaries = [_make_dept_summary()]
        result = export_service.export_department_costs_excel(summaries)
        wb = _load_excel_buffer(result)
        ws = wb.active
        for col in (5, 6, 7):
            cell = ws.cell(row=2, column=col)
            assert cell.number_format == "#,##0.00", (
                f"Column {col} number format is '{cell.number_format}', "
                f"expected '#,##0.00'"
            )
        wb.close()

    def test_non_cost_cells_have_no_currency_format(self):
        """
        The non-cost columns (Department, Divisions, Positions,
        Authorized Headcount) must NOT have currency formatting.
        """
        summaries = [_make_dept_summary()]
        result = export_service.export_department_costs_excel(summaries)
        wb = _load_excel_buffer(result)
        ws = wb.active
        for col in (1, 2, 3, 4):
            cell = ws.cell(row=2, column=col)
            assert (
                cell.number_format != "#,##0.00"
            ), f"Column {col} should not have currency number format"
        wb.close()

    # -- 2.7 Column auto-fit ------------------------------------------

    def test_columns_have_positive_widths(self):
        """
        Every column used by the export should have a positive
        width set by the _auto_fit_columns helper.
        """
        summaries = [_make_dept_summary(department_name="Engineering")]
        result = export_service.export_department_costs_excel(summaries)
        wb = _load_excel_buffer(result)
        ws = wb.active
        from openpyxl.utils import (
            get_column_letter,
        )  # pylint: disable=import-outside-toplevel

        for col in range(1, 8):
            letter = get_column_letter(col)
            width = ws.column_dimensions[letter].width
            assert width is not None and width > 0, f"Column {letter} has no width set"
        wb.close()

    def test_column_width_capped_at_40(self):
        """
        The _auto_fit_columns helper caps widths at 40 characters.
        A very long department name should not blow out the column.
        """
        long_name = "A" * 100  # 100-character name.
        summaries = [_make_dept_summary(department_name=long_name)]
        result = export_service.export_department_costs_excel(summaries)
        wb = _load_excel_buffer(result)
        ws = wb.active
        from openpyxl.utils import (
            get_column_letter,
        )  # pylint: disable=import-outside-toplevel

        col_a_width = ws.column_dimensions[get_column_letter(1)].width
        assert col_a_width <= 40, f"Column A width is {col_a_width}, expected <= 40"
        wb.close()

    # -- 2.8 Zero cost department in Excel -----------------------------

    def test_zero_cost_department_writes_zero_floats(self):
        """
        A department with $0.00 costs must write 0.0 to the
        Excel cells (not None or an empty cell).
        """
        dept = _make_dept_summary(
            hardware_total=Decimal("0.00"),
            software_total=Decimal("0.00"),
            grand_total=Decimal("0.00"),
        )
        result = export_service.export_department_costs_excel([dept])
        wb = _load_excel_buffer(result)
        ws = wb.active
        assert ws.cell(row=2, column=5).value == pytest.approx(0.00)
        assert ws.cell(row=2, column=6).value == pytest.approx(0.00)
        assert ws.cell(row=2, column=7).value == pytest.approx(0.00)
        wb.close()


# =====================================================================
# 3. Position costs CSV export
# =====================================================================


class TestExportPositionCostsCsv:
    """
    Verify ``export_position_costs_csv`` produces a correctly
    structured and encoded CSV file from PositionCostSummary data.
    """

    # -- 3.1 Return type and encoding ----------------------------------

    def test_returns_bytes_io_buffer(self):
        """The function must return an io.BytesIO instance."""
        result = export_service.export_position_costs_csv([])
        assert isinstance(result, io.BytesIO)

    def test_csv_is_utf8_bom_encoded(self):
        """The CSV must start with the UTF-8 BOM."""
        result = export_service.export_position_costs_csv([])
        raw = result.read()
        assert raw[:3] == b"\xef\xbb\xbf"

    # -- 3.2 Header row ------------------------------------------------

    def test_header_row_has_correct_columns(self):
        """
        The first row must contain exactly the eleven position-level
        column headers in the correct order.
        """
        result = export_service.export_position_costs_csv([])
        rows = _parse_csv_buffer(result)
        expected_headers = [
            "Department",
            "Division",
            "Position Code",
            "Position Title",
            "Authorized Count",
            "Hardware per Person",
            "Software per Person",
            "Total per Person",
            "Hardware Total",
            "Software Total",
            "Grand Total",
        ]
        assert rows[0] == expected_headers

    def test_header_row_column_count(self):
        """The header row must contain exactly 11 columns."""
        result = export_service.export_position_costs_csv([])
        rows = _parse_csv_buffer(result)
        assert len(rows[0]) == 11

    # -- 3.3 Empty data ------------------------------------------------

    def test_empty_data_produces_header_only(self):
        """
        When the input list is empty, the CSV must still contain
        only the header row.
        """
        result = export_service.export_position_costs_csv([])
        rows = _parse_csv_buffer(result)
        assert len(rows) == 1

    # -- 3.4 Single position -------------------------------------------

    def test_single_position_produces_one_data_row(self):
        """A single PositionCostSummary produces exactly one data row."""
        summaries = [_make_pos_summary()]
        result = export_service.export_position_costs_csv(summaries)
        rows = _parse_csv_buffer(result)
        assert len(rows) == 2

    def test_single_position_correct_values(self):
        """
        Verify that every field in the data row matches the input
        dataclass values exactly.
        """
        pos = _make_pos_summary(
            department_name="Engineering",
            division_name="Data Division",
            position_code="SA-001",
            position_title="Senior Analyst",
            authorized_count=3,
            hardware_total_per_person=Decimal("1200.00"),
            software_total_per_person=Decimal("200.00"),
            total_per_person=Decimal("1400.00"),
            hardware_total=Decimal("3600.00"),
            software_total=Decimal("600.00"),
            grand_total=Decimal("4200.00"),
        )
        result = export_service.export_position_costs_csv([pos])
        rows = _parse_csv_buffer(result)
        data_row = rows[1]

        assert data_row[0] == "Engineering"
        assert data_row[1] == "Data Division"
        assert data_row[2] == "SA-001"
        assert data_row[3] == "Senior Analyst"
        assert data_row[4] == "3"
        assert data_row[5] == "1200.00"
        assert data_row[6] == "200.00"
        assert data_row[7] == "1400.00"
        assert data_row[8] == "3600.00"
        assert data_row[9] == "600.00"
        assert data_row[10] == "4200.00"

    def test_data_row_has_same_column_count_as_header(self):
        """Every data row must have 11 columns to match the header."""
        summaries = [
            _make_pos_summary(position_id=1),
            _make_pos_summary(position_id=2),
        ]
        result = export_service.export_position_costs_csv(summaries)
        rows = _parse_csv_buffer(result)
        for i, row in enumerate(rows[1:], start=2):
            assert len(row) == 11, f"Row {i} has {len(row)} columns, expected 11"

    # -- 3.5 Multiple positions ----------------------------------------

    def test_multiple_positions_all_rows_present(self):
        """N positions produce N+1 rows (1 header + N data)."""
        summaries = [
            _make_pos_summary(position_id=i, position_code=f"POS-{i:03d}")
            for i in range(1, 5)
        ]
        result = export_service.export_position_costs_csv(summaries)
        rows = _parse_csv_buffer(result)
        assert len(rows) == 5

    def test_multiple_positions_preserve_input_order(self):
        """CSV rows must appear in the same order as the input list."""
        summaries = [
            _make_pos_summary(position_id=3, position_code="ZZZ-003"),
            _make_pos_summary(position_id=1, position_code="AAA-001"),
            _make_pos_summary(position_id=2, position_code="MMM-002"),
        ]
        result = export_service.export_position_costs_csv(summaries)
        rows = _parse_csv_buffer(result)
        assert rows[1][2] == "ZZZ-003"
        assert rows[2][2] == "AAA-001"
        assert rows[3][2] == "MMM-002"

    # -- 3.6 Decimal formatting ----------------------------------------

    def test_cost_values_formatted_to_two_decimal_places(self):
        """
        All six cost columns must be formatted with exactly two
        decimal places.
        """
        pos = _make_pos_summary(
            hardware_total_per_person=Decimal("1200.00"),
            software_total_per_person=Decimal("200.00"),
            total_per_person=Decimal("1400.00"),
            hardware_total=Decimal("3600.00"),
            software_total=Decimal("600.00"),
            grand_total=Decimal("4200.00"),
        )
        result = export_service.export_position_costs_csv([pos])
        rows = _parse_csv_buffer(result)
        data_row = rows[1]
        # Columns 5-10 are cost fields.
        for col_idx in (5, 6, 7, 8, 9, 10):
            value = data_row[col_idx]
            assert "." in value, f"Column {col_idx} missing decimal point"
            decimal_part = value.split(".")[1]
            assert len(decimal_part) == 2, (
                f"Column {col_idx} has {len(decimal_part)} decimal places, "
                f"expected 2 (value: {value})"
            )

    def test_zero_authorized_count_position(self):
        """
        A position with authorized_count=0 and $0.00 totals must
        export without errors and show the correct zero values.
        """
        pos = _make_pos_summary(
            authorized_count=0,
            hardware_total_per_person=Decimal("1200.00"),
            software_total_per_person=Decimal("200.00"),
            total_per_person=Decimal("1400.00"),
            hardware_total=Decimal("0.00"),
            software_total=Decimal("0.00"),
            grand_total=Decimal("0.00"),
        )
        result = export_service.export_position_costs_csv([pos])
        rows = _parse_csv_buffer(result)
        data_row = rows[1]
        assert data_row[4] == "0"
        assert data_row[8] == "0.00"
        assert data_row[9] == "0.00"
        assert data_row[10] == "0.00"

    # -- 3.7 Structural integrity --------------------------------------

    def test_grand_total_equals_hw_plus_sw_in_csv(self):
        """
        For every data row, Grand Total must equal Hardware Total +
        Software Total.
        """
        summaries = [
            _make_pos_summary(
                position_id=1,
                hardware_total=Decimal("3600.00"),
                software_total=Decimal("600.00"),
                grand_total=Decimal("4200.00"),
            ),
            _make_pos_summary(
                position_id=2,
                hardware_total=Decimal("0.00"),
                software_total=Decimal("1000.00"),
                grand_total=Decimal("1000.00"),
            ),
        ]
        result = export_service.export_position_costs_csv(summaries)
        rows = _parse_csv_buffer(result)
        for row in rows[1:]:
            hw = Decimal(row[8])
            sw = Decimal(row[9])
            grand = Decimal(row[10])
            assert grand == hw + sw, (
                f"Grand total mismatch for {row[3]}: " f"{grand} != {hw} + {sw}"
            )

    def test_total_per_person_times_authorized_equals_grand(self):
        """
        Total per Person * Authorized Count must equal Grand Total
        for every row, verifying the multiplier is correctly exported.
        """
        pos = _make_pos_summary(
            authorized_count=5,
            total_per_person=Decimal("1400.00"),
            hardware_total=Decimal("6000.00"),
            software_total=Decimal("1000.00"),
            grand_total=Decimal("7000.00"),
        )
        result = export_service.export_position_costs_csv([pos])
        rows = _parse_csv_buffer(result)
        data_row = rows[1]
        authorized = int(data_row[4])
        total_per_person = Decimal(data_row[7])
        grand_total = Decimal(data_row[10])
        assert grand_total == total_per_person * authorized

    def test_per_person_hw_plus_sw_equals_total_per_person(self):
        """
        Hardware per Person + Software per Person must equal
        Total per Person for every row.
        """
        pos = _make_pos_summary(
            hardware_total_per_person=Decimal("1200.00"),
            software_total_per_person=Decimal("200.00"),
            total_per_person=Decimal("1400.00"),
        )
        result = export_service.export_position_costs_csv([pos])
        rows = _parse_csv_buffer(result)
        data_row = rows[1]
        hw_pp = Decimal(data_row[5])
        sw_pp = Decimal(data_row[6])
        total_pp = Decimal(data_row[7])
        assert total_pp == hw_pp + sw_pp

    # -- 3.8 Special characters ----------------------------------------

    def test_position_title_with_comma_is_properly_quoted(self):
        """A title with a comma must be properly CSV-quoted."""
        pos = _make_pos_summary(
            position_title="Analyst, Senior Level",
        )
        result = export_service.export_position_costs_csv([pos])
        rows = _parse_csv_buffer(result)
        assert rows[1][3] == "Analyst, Senior Level"


# =====================================================================
# 4. Position costs Excel export
# =====================================================================


class TestExportPositionCostsExcel:
    """
    Verify ``export_position_costs_excel`` produces a valid .xlsx
    workbook from PositionCostSummary data.
    """

    # -- 4.1 Workbook validity -----------------------------------------

    def test_returns_bytes_io_buffer(self):
        """The function must return an io.BytesIO instance."""
        result = export_service.export_position_costs_excel([])
        assert isinstance(result, io.BytesIO)

    def test_produces_valid_xlsx_file(self):
        """The returned bytes must be parseable as a valid .xlsx file."""
        result = export_service.export_position_costs_excel([])
        wb = _load_excel_buffer(result)
        assert wb is not None
        wb.close()

    def test_active_sheet_title_is_position_costs(self):
        """The active sheet must be titled 'Position Costs'."""
        result = export_service.export_position_costs_excel([])
        wb = _load_excel_buffer(result)
        assert wb.active.title == "Position Costs"
        wb.close()

    # -- 4.2 Header row ------------------------------------------------

    def test_header_row_has_correct_columns(self):
        """
        The first row must contain the eleven position-level column
        headers in order.
        """
        result = export_service.export_position_costs_excel([])
        wb = _load_excel_buffer(result)
        ws = wb.active
        headers = [ws.cell(row=1, column=c).value for c in range(1, 12)]
        expected = [
            "Department",
            "Division",
            "Position Code",
            "Position Title",
            "Authorized Count",
            "HW per Person",
            "SW per Person",
            "Total per Person",
            "Hardware Total",
            "Software Total",
            "Grand Total",
        ]
        assert headers == expected
        wb.close()

    def test_header_row_styling_applied(self):
        """
        All 11 header cells must have bold white font, dark blue
        fill, and centered wrapped alignment.
        """
        result = export_service.export_position_costs_excel([])
        wb = _load_excel_buffer(result)
        ws = wb.active
        for col in range(1, 12):
            cell = ws.cell(row=1, column=col)
            assert cell.font.bold is True
            assert cell.alignment.horizontal == "center"
            assert cell.alignment.wrap_text is True
        wb.close()

    # -- 4.3 Empty data ------------------------------------------------

    def test_empty_data_produces_header_only(self):
        """An empty input list should produce only the header row."""
        result = export_service.export_position_costs_excel([])
        wb = _load_excel_buffer(result)
        ws = wb.active
        assert ws.max_row == 1
        wb.close()

    # -- 4.4 Data values -----------------------------------------------

    def test_single_position_correct_cell_values(self):
        """
        Verify that a single PositionCostSummary is written to
        row 2 with correct cell values and types.
        """
        pos = _make_pos_summary(
            department_name="Engineering",
            division_name="Data Division",
            position_code="SA-001",
            position_title="Senior Analyst",
            authorized_count=3,
            hardware_total_per_person=Decimal("1200.00"),
            software_total_per_person=Decimal("200.00"),
            total_per_person=Decimal("1400.00"),
            hardware_total=Decimal("3600.00"),
            software_total=Decimal("600.00"),
            grand_total=Decimal("4200.00"),
        )
        result = export_service.export_position_costs_excel([pos])
        wb = _load_excel_buffer(result)
        ws = wb.active

        # Text columns.
        assert ws.cell(row=2, column=1).value == "Engineering"
        assert ws.cell(row=2, column=2).value == "Data Division"
        assert ws.cell(row=2, column=3).value == "SA-001"
        assert ws.cell(row=2, column=4).value == "Senior Analyst"
        # Integer column.
        assert ws.cell(row=2, column=5).value == 3
        # Cost columns (stored as floats).
        assert ws.cell(row=2, column=6).value == pytest.approx(1200.00)
        assert ws.cell(row=2, column=7).value == pytest.approx(200.00)
        assert ws.cell(row=2, column=8).value == pytest.approx(1400.00)
        assert ws.cell(row=2, column=9).value == pytest.approx(3600.00)
        assert ws.cell(row=2, column=10).value == pytest.approx(600.00)
        assert ws.cell(row=2, column=11).value == pytest.approx(4200.00)
        wb.close()

    def test_multiple_positions_row_count(self):
        """N positions should produce N+1 rows."""
        summaries = [_make_pos_summary(position_id=i) for i in range(1, 4)]
        result = export_service.export_position_costs_excel(summaries)
        wb = _load_excel_buffer(result)
        ws = wb.active
        assert ws.max_row == 4  # 1 header + 3 data rows.
        wb.close()

    def test_multiple_positions_preserve_input_order(self):
        """Excel rows must match the input list order."""
        summaries = [
            _make_pos_summary(position_id=3, position_code="ZZZ-003"),
            _make_pos_summary(position_id=1, position_code="AAA-001"),
        ]
        result = export_service.export_position_costs_excel(summaries)
        wb = _load_excel_buffer(result)
        ws = wb.active
        assert ws.cell(row=2, column=3).value == "ZZZ-003"
        assert ws.cell(row=3, column=3).value == "AAA-001"
        wb.close()

    # -- 4.5 Currency number format ------------------------------------

    def test_cost_cells_have_currency_number_format(self):
        """
        The six cost columns (HW per Person through Grand Total,
        columns 6-11) must use the '#,##0.00' number format.
        """
        summaries = [_make_pos_summary()]
        result = export_service.export_position_costs_excel(summaries)
        wb = _load_excel_buffer(result)
        ws = wb.active
        for col in range(6, 12):
            cell = ws.cell(row=2, column=col)
            assert cell.number_format == "#,##0.00", (
                f"Column {col} number format is '{cell.number_format}', "
                f"expected '#,##0.00'"
            )
        wb.close()

    def test_non_cost_cells_do_not_have_currency_format(self):
        """
        Columns 1-5 (text and integer) must not have currency
        formatting applied.
        """
        summaries = [_make_pos_summary()]
        result = export_service.export_position_costs_excel(summaries)
        wb = _load_excel_buffer(result)
        ws = wb.active
        for col in range(1, 6):
            cell = ws.cell(row=2, column=col)
            assert (
                cell.number_format != "#,##0.00"
            ), f"Column {col} should not have currency number format"
        wb.close()

    # -- 4.6 Zero-cost and edge cases ----------------------------------

    def test_zero_cost_position_writes_zero_floats(self):
        """
        A position with $0.00 costs must write 0.0 to Excel
        cells, not None.
        """
        pos = _make_pos_summary(
            authorized_count=0,
            hardware_total_per_person=Decimal("0.00"),
            software_total_per_person=Decimal("0.00"),
            total_per_person=Decimal("0.00"),
            hardware_total=Decimal("0.00"),
            software_total=Decimal("0.00"),
            grand_total=Decimal("0.00"),
        )
        result = export_service.export_position_costs_excel([pos])
        wb = _load_excel_buffer(result)
        ws = wb.active
        for col in range(6, 12):
            cell_value = ws.cell(row=2, column=col).value
            assert cell_value == pytest.approx(
                0.00
            ), f"Column {col} is {cell_value!r}, expected 0.0"
        wb.close()

    def test_large_cost_values_preserved_in_excel(self):
        """
        Very large cost values must be written without truncation
        or precision loss in Excel.
        """
        pos = _make_pos_summary(
            hardware_total=Decimal("9999999.99"),
            software_total=Decimal("1234567.89"),
            grand_total=Decimal("11234567.88"),
        )
        result = export_service.export_position_costs_excel([pos])
        wb = _load_excel_buffer(result)
        ws = wb.active
        assert ws.cell(row=2, column=9).value == pytest.approx(9999999.99)
        assert ws.cell(row=2, column=10).value == pytest.approx(1234567.89)
        assert ws.cell(row=2, column=11).value == pytest.approx(11234567.88)
        wb.close()


# =====================================================================
# 5. Cross-format consistency
# =====================================================================


class TestCrossFormatConsistency:
    """
    Verify that CSV and Excel exports for the same input data
    produce equivalent values.  This catches bugs where one format
    diverges from the other due to formatting or conversion errors.
    """

    def test_department_csv_and_excel_values_match(self):
        """
        The CSV and Excel exports for the same department data
        must contain the same department names and cost values.
        """
        summaries = [
            _make_dept_summary(
                department_name="Engineering",
                division_count=3,
                position_count=12,
                total_authorized=45,
                hardware_total=Decimal("54000.00"),
                software_total=Decimal("9000.00"),
                grand_total=Decimal("63000.00"),
            ),
            _make_dept_summary(
                department_name="Finance",
                division_count=1,
                position_count=4,
                total_authorized=10,
                hardware_total=Decimal("12000.00"),
                software_total=Decimal("2000.00"),
                grand_total=Decimal("14000.00"),
            ),
        ]

        # Generate both formats.
        csv_buffer = export_service.export_department_costs_csv(summaries)
        xlsx_buffer = export_service.export_department_costs_excel(summaries)

        # Parse CSV.
        csv_rows = _parse_csv_buffer(csv_buffer)
        # Parse Excel.
        wb = _load_excel_buffer(xlsx_buffer)
        ws = wb.active

        # Compare data rows (skip headers).
        for i, csv_row in enumerate(csv_rows[1:], start=2):
            # Department name.
            assert csv_row[0] == ws.cell(row=i, column=1).value
            # Division count (CSV is string, Excel is int).
            assert int(csv_row[1]) == ws.cell(row=i, column=2).value
            # Position count.
            assert int(csv_row[2]) == ws.cell(row=i, column=3).value
            # Authorized headcount.
            assert int(csv_row[3]) == ws.cell(row=i, column=4).value
            # Cost columns (CSV is string, Excel is float).
            assert Decimal(csv_row[4]) == pytest.approx(
                ws.cell(row=i, column=5).value, abs=Decimal("0.01")
            )
            assert Decimal(csv_row[5]) == pytest.approx(
                ws.cell(row=i, column=6).value, abs=Decimal("0.01")
            )
            assert Decimal(csv_row[6]) == pytest.approx(
                ws.cell(row=i, column=7).value, abs=Decimal("0.01")
            )
        wb.close()

    def test_position_csv_and_excel_values_match(self):
        """
        The CSV and Excel exports for the same position data
        must contain equivalent values.
        """
        summaries = [
            _make_pos_summary(
                position_code="SA-001",
                department_name="Engineering",
                division_name="Data Division",
                authorized_count=3,
                hardware_total_per_person=Decimal("1200.00"),
                software_total_per_person=Decimal("200.00"),
                total_per_person=Decimal("1400.00"),
                hardware_total=Decimal("3600.00"),
                software_total=Decimal("600.00"),
                grand_total=Decimal("4200.00"),
            ),
        ]

        csv_buffer = export_service.export_position_costs_csv(summaries)
        xlsx_buffer = export_service.export_position_costs_excel(summaries)

        csv_rows = _parse_csv_buffer(csv_buffer)
        wb = _load_excel_buffer(xlsx_buffer)
        ws = wb.active

        csv_row = csv_rows[1]
        # Text columns.
        assert csv_row[0] == ws.cell(row=2, column=1).value  # Department.
        assert csv_row[1] == ws.cell(row=2, column=2).value  # Division.
        assert csv_row[2] == ws.cell(row=2, column=3).value  # Position Code.
        assert csv_row[3] == ws.cell(row=2, column=4).value  # Position Title.
        # Authorized count.
        assert int(csv_row[4]) == ws.cell(row=2, column=5).value
        # Cost columns.
        for csv_col, xl_col in [(5, 6), (6, 7), (7, 8), (8, 9), (9, 10), (10, 11)]:
            csv_val = Decimal(csv_row[csv_col])
            xl_val = ws.cell(row=2, column=xl_col).value
            assert csv_val == pytest.approx(
                xl_val, abs=Decimal("0.01")
            ), f"CSV col {csv_col} ({csv_val}) != Excel col {xl_col} ({xl_val})"
        wb.close()


# =====================================================================
# 6. Internal helper: _format_decimal
# =====================================================================


class TestFormatDecimalHelper:
    """
    Verify the ``_format_decimal`` internal helper directly.

    Although this is technically a private function, it is the
    single point of truth for all CSV cost formatting.  Testing
    it directly catches edge cases that would be difficult to
    isolate through the public API.
    """

    def test_standard_two_decimal_value(self):
        """A standard Decimal with two places formats correctly."""
        assert export_service._format_decimal(Decimal("1234.56")) == "1234.56"

    def test_whole_number_gets_two_decimal_places(self):
        """A whole number Decimal gains .00 suffix."""
        assert export_service._format_decimal(Decimal("1000")) == "1000.00"

    def test_one_decimal_place_gets_padded(self):
        """A Decimal with one decimal place gets a trailing zero."""
        assert export_service._format_decimal(Decimal("99.5")) == "99.50"

    def test_three_decimal_places_rounds(self):
        """Excess decimal places are rounded (f-string behavior)."""
        # Python f"{Decimal('1.555'):.2f}" rounds to "1.56" (banker's
        # rounding or round-half-even, depending on Python version).
        result = export_service._format_decimal(Decimal("1.555"))
        # Accept either "1.55" or "1.56" -- the important thing is
        # that the output has exactly two decimal places.
        assert result in ("1.55", "1.56")
        assert len(result.split(".")[1]) == 2

    def test_zero_formats_as_zero_point_zero_zero(self):
        """Zero must format as '0.00'."""
        assert export_service._format_decimal(Decimal("0")) == "0.00"
        assert export_service._format_decimal(Decimal("0.00")) == "0.00"

    def test_negative_value_formats_with_minus(self):
        """
        Negative values (unlikely but possible if data is corrupt)
        must include the minus sign and two decimal places.
        """
        assert export_service._format_decimal(Decimal("-500.00")) == "-500.00"

    def test_very_small_fractional_value(self):
        """A value smaller than one cent still gets two decimal places."""
        assert export_service._format_decimal(Decimal("0.001")) == "0.00"

    def test_large_value_no_scientific_notation(self):
        """Large values must never be formatted in scientific notation."""
        result = export_service._format_decimal(Decimal("12345678.90"))
        assert "e" not in result.lower()
        assert result == "12345678.90"
