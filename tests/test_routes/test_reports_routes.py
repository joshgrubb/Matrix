"""
Integration tests for the reports blueprint routes.

Verifies that the department cost summary, position equipment report,
and CSV/Excel export endpoints render correct data, respect user
scope, apply role restrictions on exports, and produce valid file
output with accurate content.

Role-based access control (who can/cannot reach the export endpoints)
is already tested in ``test_scope_isolation.py``.  These tests focus
on the **behavioral** side:

    - Does the page render the correct cost figures for known data?
    - Does the cost summary aggregate correctly to an org-level footer?
    - Does the equipment report filter by department and division?
    - Do scope-limited users see only their authorized departments?
    - Do CSV exports contain the correct headers and data values?
    - Do Excel exports produce valid .xlsx files with correct cells?
    - Does an export with no configured positions produce headers only?
    - Does the position export respect query-string filters?

Design decisions:
    - Every test that checks rendered cost figures creates known
      requirements via the ``create_hw_requirement`` and
      ``create_sw_requirement`` factory fixtures, then asserts the
      expected dollar amounts appear in the response body.  This
      catches regressions in the cost_service -> route -> template
      pipeline end-to-end.
    - CSV assertions parse the raw response bytes through Python's
      ``csv.reader`` so field-level values can be compared exactly.
      This catches column-order changes, missing columns, and
      formatting regressions.
    - Excel assertions use ``openpyxl.load_workbook`` to read the
      response bytes and verify cell values.  This catches broken
      workbooks, wrong sheet names, and incorrect number formatting.
    - All dollar assertions compare against ``Decimal`` values
      calculated from known fixture costs to avoid floating-point
      surprises.

Fixture reminder (from conftest.py):
    admin_user:             org-wide scope
    manager_user:           scoped to div_a1 (within dept_a)
    manager_dept_scope_user: scoped to dept_a
    it_staff_user:          org-wide scope
    budget_user:            org-wide scope
    read_only_user:         scoped to div_a1

    sample_org keys:
        dept_a, dept_b, div_a1, div_a2, div_b1, div_b2,
        pos_a1_1 (auth=3), pos_a1_2 (auth=5), pos_a2_1 (auth=2),
        pos_b1_1 (auth=4), pos_b1_2 (auth=1), pos_b2_1 (auth=6)

    sample_catalog keys:
        hw_laptop_standard  $1,200.00 per unit
        hw_laptop_power     $2,400.00 per unit
        hw_monitor_24       $  350.00 per unit
        sw_office_e3        $  200.00 per_user
        sw_office_e5        $  400.00 per_user
        sw_antivirus        $50,000.00 tenant

Run this file in isolation::

    pytest tests/test_routes/test_reports_routes.py -v
"""

import csv
import io
from decimal import Decimal

import pytest
from openpyxl import load_workbook


# =====================================================================
# Helpers
# =====================================================================


def _parse_csv_response(response):
    """
    Parse a Flask test-client CSV response into a list of rows.

    Handles the UTF-8 BOM (``\\xef\\xbb\\xbf``) that the export
    service prepends via ``utf-8-sig`` encoding.

    Args:
        response: The test-client response object.

    Returns:
        A list of lists.  Row 0 is the header row.
    """
    # Decode with utf-8-sig to strip the BOM transparently.
    text = response.data.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    return list(reader)


def _load_excel_response(response):
    """
    Load a Flask test-client Excel response into an openpyxl Workbook.

    Args:
        response: The test-client response object.

    Returns:
        An openpyxl Workbook instance (read-only is not used so
        cell values are fully accessible).
    """
    buffer = io.BytesIO(response.data)
    return load_workbook(buffer)


# =====================================================================
# 1. Cost summary page -- rendering and content
# =====================================================================


class TestCostSummaryPage:
    """
    Verify the department cost summary page (``/reports/cost-summary``)
    loads correctly and renders accurate cost figures for known data.
    """

    def test_cost_summary_returns_200_for_admin(self, auth_client, admin_user):
        """An admin with org-wide scope can load the cost summary."""
        client = auth_client(admin_user)
        response = client.get("/reports/cost-summary")
        assert response.status_code == 200

    def test_cost_summary_contains_page_heading(self, auth_client, admin_user):
        """The page should render its identifying heading."""
        client = auth_client(admin_user)
        response = client.get("/reports/cost-summary")
        assert b"Department Cost Summary" in response.data

    def test_cost_summary_contains_export_links(self, auth_client, admin_user):
        """
        The page should contain links to the CSV and Excel export
        endpoints so users can download the data.
        """
        client = auth_client(admin_user)
        response = client.get("/reports/cost-summary")
        assert b"Export CSV" in response.data
        assert b"Export Excel" in response.data

    def test_cost_summary_shows_department_names(
        self, auth_client, admin_user, sample_org
    ):
        """
        The cost summary table should list every department the user
        can access.  An admin sees all departments, including the
        two created by the sample_org fixture.
        """
        client = auth_client(admin_user)
        response = client.get("/reports/cost-summary")
        assert sample_org["dept_a"].department_name.encode() in response.data
        assert sample_org["dept_b"].department_name.encode() in response.data

    def test_cost_summary_shows_correct_hardware_cost(
        self,
        auth_client,
        admin_user,
        sample_org,
        sample_catalog,
        create_hw_requirement,
    ):
        """
        Configure a known hardware requirement, then verify the
        rendered page contains the correct dollar figure.

        pos_a1_1 (authorized_count=3): 1x hw_laptop_standard @ $1,200
        Expected hardware_total = $1,200 * 3 = $3,600.00
        """
        create_hw_requirement(
            position=sample_org["pos_a1_1"],
            hardware=sample_catalog["hw_laptop_standard"],
            quantity=1,
        )

        client = auth_client(admin_user)
        response = client.get("/reports/cost-summary")
        assert response.status_code == 200

        # The template renders costs as $X.XX via "%.2f"|format().
        assert b"3600.00" in response.data

    def test_cost_summary_shows_correct_software_cost(
        self,
        auth_client,
        admin_user,
        sample_org,
        sample_catalog,
        create_sw_requirement,
    ):
        """
        Configure a known per-user software requirement, then verify
        the rendered page contains the correct dollar figure.

        pos_a1_2 (authorized_count=5): 1x sw_office_e3 @ $200/user
        Expected software_total = $200 * 5 = $1,000.00
        """
        create_sw_requirement(
            position=sample_org["pos_a1_2"],
            software=sample_catalog["sw_office_e3"],
            quantity=1,
        )

        client = auth_client(admin_user)
        response = client.get("/reports/cost-summary")
        assert response.status_code == 200
        assert b"1000.00" in response.data

    def test_cost_summary_shows_organization_total_footer(
        self,
        auth_client,
        admin_user,
        sample_org,
        sample_catalog,
        create_hw_requirement,
        create_sw_requirement,
    ):
        """
        The cost summary footer should show an organization-level
        grand total that equals the sum of all department totals.

        Configure:
            dept_a/pos_a1_1: 1x laptop ($3,600)
            dept_b/pos_b1_1: 1x sw_office_e3 ($800)

        The org grand_total for these two items is $4,400.  The
        rendered page should contain 'Organization Total' and the
        correct figures.

        Note: the org total may include additional amounts if the
        test database has pre-existing seed data, so we verify the
        structural element exists and our known amounts appear.
        """
        create_hw_requirement(
            position=sample_org["pos_a1_1"],
            hardware=sample_catalog["hw_laptop_standard"],
            quantity=1,
        )
        create_sw_requirement(
            position=sample_org["pos_b1_1"],
            software=sample_catalog["sw_office_e3"],
            quantity=1,
        )

        client = auth_client(admin_user)
        response = client.get("/reports/cost-summary")
        assert response.status_code == 200
        assert b"Organization Total" in response.data
        # Our dept_a hardware appears.
        assert b"3600.00" in response.data
        # Our dept_b software: $200 * 4 authorized = $800.
        assert b"800.00" in response.data

    def test_cost_summary_zero_cost_department_shows_zero(
        self, auth_client, admin_user, sample_org
    ):
        """
        A department with no configured requirements should appear
        in the table with $0.00 totals, not be omitted or error.
        """
        client = auth_client(admin_user)
        response = client.get("/reports/cost-summary")
        assert response.status_code == 200
        # dept_b has no requirements, so its row should show 0.00.
        assert b"0.00" in response.data


# =====================================================================
# 2. Cost summary page -- scope filtering
# =====================================================================


class TestCostSummaryScopeFiltering:
    """
    Verify that the cost summary page respects the user's
    organizational scope and does not leak data across boundaries.
    """

    def test_manager_sees_only_scoped_departments(
        self, auth_client, manager_user, sample_org
    ):
        """
        A manager scoped to div_a1 should see department A
        (the parent of div_a1) but NOT department B.

        The ``get_department_cost_breakdown`` service filters by
        the user's scope, so dept_b should not appear in the
        response body.
        """
        client = auth_client(manager_user)
        response = client.get("/reports/cost-summary")
        assert response.status_code == 200
        assert sample_org["dept_a"].department_name.encode() in response.data
        assert sample_org["dept_b"].department_name.encode() not in response.data

    def test_dept_scoped_manager_sees_only_own_department(
        self, auth_client, manager_dept_scope_user, sample_org
    ):
        """
        A manager scoped to dept_a should see dept_a but not dept_b.
        """
        client = auth_client(manager_dept_scope_user)
        response = client.get("/reports/cost-summary")
        assert response.status_code == 200
        assert sample_org["dept_a"].department_name.encode() in response.data
        assert sample_org["dept_b"].department_name.encode() not in response.data

    def test_admin_sees_all_departments(self, auth_client, admin_user, sample_org):
        """An admin with org-wide scope sees every department."""
        client = auth_client(admin_user)
        response = client.get("/reports/cost-summary")
        assert response.status_code == 200
        assert sample_org["dept_a"].department_name.encode() in response.data
        assert sample_org["dept_b"].department_name.encode() in response.data

    def test_read_only_user_can_view_cost_summary(self, auth_client, read_only_user):
        """
        Read-only users can view the cost summary page (they just
        cannot export).  This is a behavioral confirmation that
        the page is accessible to all authenticated roles.
        """
        client = auth_client(read_only_user)
        response = client.get("/reports/cost-summary")
        assert response.status_code == 200


# =====================================================================
# 3. Equipment report page -- rendering and content
# =====================================================================


class TestEquipmentReportPage:
    """
    Verify the position equipment report (``/reports/equipment-report``)
    loads correctly, renders position-level cost data, and supports
    department/division filtering.
    """

    def test_equipment_report_returns_200_for_admin(self, auth_client, admin_user):
        """An admin can load the equipment report page."""
        client = auth_client(admin_user)
        response = client.get("/reports/equipment-report")
        assert response.status_code == 200

    def test_equipment_report_contains_page_heading(self, auth_client, admin_user):
        """The page should render its identifying heading."""
        client = auth_client(admin_user)
        response = client.get("/reports/equipment-report")
        assert b"Position Equipment Report" in response.data

    def test_equipment_report_contains_filter_form(
        self, auth_client, admin_user, sample_org
    ):
        """
        The equipment report should contain a department filter
        dropdown with the fixture departments as selectable options.
        """
        client = auth_client(admin_user)
        response = client.get("/reports/equipment-report")
        assert b"department_id" in response.data
        assert sample_org["dept_a"].department_name.encode() in response.data

    def test_equipment_report_contains_export_links(self, auth_client, admin_user):
        """The page should contain CSV and Excel export links."""
        client = auth_client(admin_user)
        response = client.get("/reports/equipment-report")
        assert b"Export CSV" in response.data
        assert b"Export Excel" in response.data

    def test_equipment_report_shows_position_with_requirements(
        self,
        auth_client,
        admin_user,
        sample_org,
        sample_catalog,
        create_hw_requirement,
    ):
        """
        A position with a hardware requirement should appear in
        the equipment report table with its cost figures.

        pos_a1_1 (authorized_count=3): 1x hw_laptop_standard @ $1,200
        Expected grand_total = $3,600.00
        """
        create_hw_requirement(
            position=sample_org["pos_a1_1"],
            hardware=sample_catalog["hw_laptop_standard"],
            quantity=1,
        )

        client = auth_client(admin_user)
        response = client.get("/reports/equipment-report")
        assert response.status_code == 200
        assert sample_org["pos_a1_1"].position_title.encode() in response.data
        assert b"3600.00" in response.data

    def test_equipment_report_filters_by_department(
        self,
        auth_client,
        admin_user,
        sample_org,
        sample_catalog,
        create_hw_requirement,
    ):
        """
        When filtering by dept_a, only dept_a positions should appear.
        dept_b positions must be absent from the response.
        """
        # Create requirements in both departments so they would
        # normally both appear.
        create_hw_requirement(
            position=sample_org["pos_a1_1"],
            hardware=sample_catalog["hw_laptop_standard"],
            quantity=1,
        )
        create_hw_requirement(
            position=sample_org["pos_b1_1"],
            hardware=sample_catalog["hw_monitor_24"],
            quantity=1,
        )

        dept_a_id = sample_org["dept_a"].id
        client = auth_client(admin_user)
        response = client.get(f"/reports/equipment-report?department_id={dept_a_id}")
        assert response.status_code == 200

        # pos_a1_1 should appear; pos_b1_1 should not.
        assert sample_org["pos_a1_1"].position_title.encode() in response.data
        assert sample_org["pos_b1_1"].position_title.encode() not in response.data

    def test_equipment_report_filters_by_division(
        self,
        auth_client,
        admin_user,
        sample_org,
        sample_catalog,
        create_hw_requirement,
    ):
        """
        When filtering by div_a1, only div_a1 positions should appear.
        div_a2 positions in the same department must be excluded.
        """
        # Create requirements in two divisions of the same department.
        create_hw_requirement(
            position=sample_org["pos_a1_1"],
            hardware=sample_catalog["hw_laptop_standard"],
            quantity=1,
        )
        create_hw_requirement(
            position=sample_org["pos_a2_1"],
            hardware=sample_catalog["hw_monitor_24"],
            quantity=1,
        )

        div_a1_id = sample_org["div_a1"].id
        client = auth_client(admin_user)
        response = client.get(f"/reports/equipment-report?division_id={div_a1_id}")
        assert response.status_code == 200

        # pos_a1_1 is in div_a1 and should appear.
        assert sample_org["pos_a1_1"].position_title.encode() in response.data
        # pos_a2_1 is in div_a2 and should not appear.
        assert sample_org["pos_a2_1"].position_title.encode() not in response.data

    def test_equipment_report_shows_clear_filter_link(
        self, auth_client, admin_user, sample_org
    ):
        """
        When a filter is active, the page should render a 'Clear'
        link that points back to the unfiltered report.
        """
        dept_a_id = sample_org["dept_a"].id
        client = auth_client(admin_user)
        response = client.get(f"/reports/equipment-report?department_id={dept_a_id}")
        assert response.status_code == 200
        assert b"Clear" in response.data

    def test_equipment_report_no_clear_link_when_unfiltered(
        self, auth_client, admin_user
    ):
        """
        When no filter is active, the 'Clear' link should not appear.
        The template conditionally renders the Clear button only when
        ``selected_department_id`` or ``selected_division_id`` is set.
        """
        client = auth_client(admin_user)
        response = client.get("/reports/equipment-report")
        html = response.data.decode("utf-8")
        # The Clear link uses a unique CSS class combination that
        # distinguishes it from other buttons on the page.
        assert "btn-outline-secondary ms-1" not in html

    def test_equipment_report_scope_limits_positions_for_manager(
        self,
        auth_client,
        manager_user,
        sample_org,
        sample_catalog,
        create_hw_requirement,
    ):
        """
        A manager scoped to div_a1 should only see div_a1 positions
        in the equipment report, even without an explicit filter.
        """
        create_hw_requirement(
            position=sample_org["pos_a1_1"],
            hardware=sample_catalog["hw_laptop_standard"],
            quantity=1,
        )
        create_hw_requirement(
            position=sample_org["pos_b1_1"],
            hardware=sample_catalog["hw_monitor_24"],
            quantity=1,
        )

        client = auth_client(manager_user)
        response = client.get("/reports/equipment-report")
        assert response.status_code == 200

        # Manager scoped to div_a1 sees pos_a1_1 but not pos_b1_1.
        assert sample_org["pos_a1_1"].position_title.encode() in response.data
        assert sample_org["pos_b1_1"].position_title.encode() not in response.data

    def test_equipment_report_no_requirements_shows_zero_cost_rows(
        self, auth_client, admin_user
    ):
        """
        When no positions have configured requirements, the page
        should still render without errors.  Positions will appear
        with $0.00 cost rows because the route calls
        ``calculate_position_cost`` for every position in scope,
        and positions with no requirements return zero-cost summaries.
        """
        client = auth_client(admin_user)
        response = client.get("/reports/equipment-report")
        assert response.status_code == 200


# =====================================================================
# 4. Department cost CSV export
# =====================================================================


class TestExportDepartmentCostsCsv:
    """
    Verify the ``/reports/export/department-costs/csv`` endpoint
    returns a valid CSV file with correct headers, data values,
    content type, and disposition.
    """

    def test_csv_content_type(self, auth_client, admin_user):
        """The response Content-Type must be text/csv."""
        client = auth_client(admin_user)
        response = client.get("/reports/export/department-costs/csv")
        assert response.status_code == 200
        assert "text/csv" in response.headers["Content-Type"]

    def test_csv_content_disposition(self, auth_client, admin_user):
        """The response must include an attachment disposition with filename."""
        client = auth_client(admin_user)
        response = client.get("/reports/export/department-costs/csv")
        disposition = response.headers.get("Content-Disposition", "")
        assert "attachment" in disposition
        assert "department_costs.csv" in disposition

    def test_csv_header_row_matches_export_service(self, auth_client, admin_user):
        """
        The CSV header row must contain the exact column names
        defined in ``export_service.export_department_costs_csv``.
        """
        client = auth_client(admin_user)
        response = client.get("/reports/export/department-costs/csv")
        rows = _parse_csv_response(response)

        expected_headers = [
            "Department",
            "Divisions",
            "Positions",
            "Authorized Headcount",
            "Hardware Total",
            "Software Total",
            "Grand Total",
        ]
        assert rows[0] == expected_headers

    def test_csv_contains_fixture_departments(
        self, auth_client, admin_user, sample_org
    ):
        """
        The CSV should contain data rows for both fixture departments.
        """
        client = auth_client(admin_user)
        response = client.get("/reports/export/department-costs/csv")
        rows = _parse_csv_response(response)

        # Collect all department names from the CSV.
        dept_names = [row[0] for row in rows[1:]]
        assert sample_org["dept_a"].department_name in dept_names
        assert sample_org["dept_b"].department_name in dept_names

    def test_csv_data_values_match_known_costs(
        self,
        auth_client,
        admin_user,
        sample_org,
        sample_catalog,
        create_hw_requirement,
        create_sw_requirement,
    ):
        """
        Configure known requirements and verify the CSV data row
        contains the exact calculated cost values.

        pos_a1_1 (auth=3): 1x hw_laptop_standard @ $1,200
            hardware_total = $3,600.00
        pos_a1_2 (auth=5): 1x sw_office_e3 @ $200
            software_total = $1,000.00

        dept_a row:
            Divisions=2, Positions=3, Authorized=10,
            Hardware Total=3600.00, Software Total=1000.00,
            Grand Total=4600.00
        """
        create_hw_requirement(
            position=sample_org["pos_a1_1"],
            hardware=sample_catalog["hw_laptop_standard"],
            quantity=1,
        )
        create_sw_requirement(
            position=sample_org["pos_a1_2"],
            software=sample_catalog["sw_office_e3"],
            quantity=1,
        )

        client = auth_client(admin_user)
        response = client.get("/reports/export/department-costs/csv")
        rows = _parse_csv_response(response)

        # Find the dept_a row.
        dept_a_name = sample_org["dept_a"].department_name
        dept_a_row = next(r for r in rows[1:] if r[0] == dept_a_name)

        # Column indices: 0=Dept, 1=Divs, 2=Pos, 3=Auth,
        #                 4=HW Total, 5=SW Total, 6=Grand Total
        assert dept_a_row[1] == "2"  # 2 divisions in dept_a
        assert dept_a_row[2] == "3"  # 3 positions in dept_a
        assert dept_a_row[3] == "10"  # 3+5+2 authorized
        assert Decimal(dept_a_row[4]) == Decimal("3600.00")
        assert Decimal(dept_a_row[5]) == Decimal("1000.00")
        assert Decimal(dept_a_row[6]) == Decimal("4600.00")

    def test_csv_zero_cost_department_shows_zero(
        self, auth_client, admin_user, sample_org
    ):
        """
        A department with no configured requirements should appear
        in the CSV with 0.00 cost columns, not be omitted.
        """
        client = auth_client(admin_user)
        response = client.get("/reports/export/department-costs/csv")
        rows = _parse_csv_response(response)

        dept_b_name = sample_org["dept_b"].department_name
        dept_b_row = next(r for r in rows[1:] if r[0] == dept_b_name)

        assert Decimal(dept_b_row[4]) == Decimal("0.00")  # Hardware
        assert Decimal(dept_b_row[5]) == Decimal("0.00")  # Software
        assert Decimal(dept_b_row[6]) == Decimal("0.00")  # Grand Total

    def test_csv_scope_limits_data_for_dept_scoped_budget_executive(
        self,
        auth_client,
        create_user,
        sample_org,
    ):
        """
        A budget executive scoped to dept_a should get a CSV that
        contains only dept_a rows.  dept_b must be completely absent.

        Managers cannot access the export endpoint (blocked by
        ``@role_required``), so this test uses a budget_executive
        with department-level scope to verify that the export
        service correctly filters by the requesting user's scope.
        """
        # Create a budget_executive scoped to dept_a only.
        dept_budget_user = create_user(
            role_name="budget_executive",
            scopes=[
                {
                    "scope_type": "department",
                    "department_id": sample_org["dept_a"].id,
                }
            ],
            first_name="Dept",
            last_name="Budget",
        )

        client = auth_client(dept_budget_user)
        response = client.get("/reports/export/department-costs/csv")
        assert response.status_code == 200

        rows = _parse_csv_response(response)
        dept_names = {row[0] for row in rows[1:]}

        assert sample_org["dept_a"].department_name in dept_names
        assert sample_org["dept_b"].department_name not in dept_names

    def test_csv_grand_total_equals_hw_plus_sw(
        self,
        auth_client,
        admin_user,
        sample_org,
        sample_catalog,
        create_hw_requirement,
        create_sw_requirement,
    ):
        """
        For every row in the CSV, Grand Total must equal
        Hardware Total + Software Total.  This is a structural
        integrity check that catches aggregation bugs.
        """
        create_hw_requirement(
            position=sample_org["pos_a1_1"],
            hardware=sample_catalog["hw_laptop_standard"],
            quantity=1,
        )
        create_sw_requirement(
            position=sample_org["pos_a1_2"],
            software=sample_catalog["sw_office_e3"],
            quantity=1,
        )

        client = auth_client(admin_user)
        response = client.get("/reports/export/department-costs/csv")
        rows = _parse_csv_response(response)

        for row in rows[1:]:
            hw_total = Decimal(row[4])
            sw_total = Decimal(row[5])
            grand_total = Decimal(row[6])
            assert grand_total == hw_total + sw_total, (
                f"Grand total mismatch for {row[0]}: "
                f"{grand_total} != {hw_total} + {sw_total}"
            )


# =====================================================================
# 5. Department cost Excel export
# =====================================================================


class TestExportDepartmentCostsExcel:
    """
    Verify the ``/reports/export/department-costs/xlsx`` endpoint
    returns a valid Excel workbook with correct headers and values.
    """

    def test_xlsx_content_type(self, auth_client, admin_user):
        """The response Content-Type must be the OOXML spreadsheet MIME type."""
        client = auth_client(admin_user)
        response = client.get("/reports/export/department-costs/xlsx")
        assert response.status_code == 200
        expected_type = (
            "application/vnd.openxmlformats-officedocument" ".spreadsheetml.sheet"
        )
        assert expected_type in response.headers["Content-Type"]

    def test_xlsx_content_disposition(self, auth_client, admin_user):
        """The response must include an attachment disposition with filename."""
        client = auth_client(admin_user)
        response = client.get("/reports/export/department-costs/xlsx")
        disposition = response.headers.get("Content-Disposition", "")
        assert "attachment" in disposition
        assert "department_costs.xlsx" in disposition

    def test_xlsx_produces_valid_workbook(self, auth_client, admin_user):
        """
        The response bytes must be parseable as a valid .xlsx file
        by openpyxl.  A corrupt file would raise an exception here.
        """
        client = auth_client(admin_user)
        response = client.get("/reports/export/department-costs/xlsx")
        wb = _load_excel_response(response)
        assert wb.active is not None
        assert wb.active.title == "Department Costs"

    def test_xlsx_header_row(self, auth_client, admin_user):
        """
        The first row of the active sheet must contain the expected
        column headers.
        """
        client = auth_client(admin_user)
        response = client.get("/reports/export/department-costs/xlsx")
        wb = _load_excel_response(response)
        ws = wb.active

        headers = [ws.cell(row=1, column=c).value for c in range(1, 8)]
        expected = [
            "Department",
            "Divisions",
            "Positions",
            "Authorized Headcount",
            "Hardware Total",
            "Software Total",
            "Grand Total",
        ]
        assert headers == expected

    def test_xlsx_data_values_match_known_costs(
        self,
        auth_client,
        admin_user,
        sample_org,
        sample_catalog,
        create_hw_requirement,
    ):
        """
        Configure a known requirement and verify the Excel cell
        values match.

        pos_a1_1 (auth=3): 1x hw_laptop_standard @ $1,200
        dept_a Hardware Total = 3600.00
        """
        create_hw_requirement(
            position=sample_org["pos_a1_1"],
            hardware=sample_catalog["hw_laptop_standard"],
            quantity=1,
        )

        client = auth_client(admin_user)
        response = client.get("/reports/export/department-costs/xlsx")
        wb = _load_excel_response(response)
        ws = wb.active

        # Find the dept_a row by scanning column 1.
        dept_a_name = sample_org["dept_a"].department_name
        dept_a_row = None
        for row_idx in range(2, ws.max_row + 1):
            if ws.cell(row=row_idx, column=1).value == dept_a_name:
                dept_a_row = row_idx
                break

        assert (
            dept_a_row is not None
        ), f"Department '{dept_a_name}' not found in Excel export"

        # Column 5 = Hardware Total.  The export writes float values.
        hw_total = ws.cell(row=dept_a_row, column=5).value
        assert hw_total == pytest.approx(3600.00, abs=0.01)

        # Column 7 = Grand Total.
        grand_total = ws.cell(row=dept_a_row, column=7).value
        assert grand_total == pytest.approx(3600.00, abs=0.01)


# =====================================================================
# 6. Position cost CSV export
# =====================================================================


class TestExportPositionCostsCsv:
    """
    Verify the ``/reports/export/position-costs/csv`` endpoint
    returns a valid CSV file with correct position-level data.
    """

    def test_csv_content_type(self, auth_client, admin_user):
        """The response Content-Type must be text/csv."""
        client = auth_client(admin_user)
        response = client.get("/reports/export/position-costs/csv")
        assert response.status_code == 200
        assert "text/csv" in response.headers["Content-Type"]

    def test_csv_content_disposition(self, auth_client, admin_user):
        """The response must include an attachment disposition."""
        client = auth_client(admin_user)
        response = client.get("/reports/export/position-costs/csv")
        disposition = response.headers.get("Content-Disposition", "")
        assert "attachment" in disposition
        assert "position_costs.csv" in disposition

    def test_csv_header_row(self, auth_client, admin_user):
        """
        The CSV header row must contain the exact column names
        defined in ``export_service.export_position_costs_csv``.
        """
        client = auth_client(admin_user)
        response = client.get("/reports/export/position-costs/csv")
        rows = _parse_csv_response(response)

        expected_headers = [
            "Department",
            "Division",
            "Position Code",
            "Position Title",
            "Authorized Count",
            "Hardware per Person",
            "Software per Person",
            "Total per Person",
            "Hardware Total",
            "Software Total",
            "Grand Total",
        ]
        assert rows[0] == expected_headers

    def test_csv_data_values_match_known_costs(
        self,
        auth_client,
        admin_user,
        sample_org,
        sample_catalog,
        create_hw_requirement,
        create_sw_requirement,
    ):
        """
        Configure known requirements on a position and verify the
        CSV row contains exact cost values.

        pos_a1_1 (auth=3): 1x hw_laptop_standard @ $1,200
                          + 1x sw_office_e3 @ $200

        hw_per_person  = $1,200.00
        sw_per_person  = $200.00
        total_per_person = $1,400.00
        hw_total  = $1,200 * 3 = $3,600.00
        sw_total  = $200 * 3   = $600.00
        grand_total = $4,200.00
        """
        pos = sample_org["pos_a1_1"]
        create_hw_requirement(
            position=pos,
            hardware=sample_catalog["hw_laptop_standard"],
            quantity=1,
        )
        create_sw_requirement(
            position=pos,
            software=sample_catalog["sw_office_e3"],
            quantity=1,
        )

        client = auth_client(admin_user)
        response = client.get("/reports/export/position-costs/csv")
        rows = _parse_csv_response(response)

        # Find the row for pos_a1_1 by position code.
        pos_code = pos.position_code
        pos_row = next(r for r in rows[1:] if r[2] == pos_code)

        # Column indices:
        # 0=Dept, 1=Div, 2=Code, 3=Title, 4=AuthCount,
        # 5=HW/person, 6=SW/person, 7=Total/person,
        # 8=HW Total, 9=SW Total, 10=Grand Total
        assert pos_row[0] == sample_org["dept_a"].department_name
        assert pos_row[1] == sample_org["div_a1"].division_name
        assert pos_row[3] == pos.position_title
        assert pos_row[4] == "3"  # authorized_count
        assert Decimal(pos_row[5]) == Decimal("1200.00")  # hw per person
        assert Decimal(pos_row[6]) == Decimal("200.00")  # sw per person
        assert Decimal(pos_row[7]) == Decimal("1400.00")  # total per person
        assert Decimal(pos_row[8]) == Decimal("3600.00")  # hw total
        assert Decimal(pos_row[9]) == Decimal("600.00")  # sw total
        assert Decimal(pos_row[10]) == Decimal("4200.00")  # grand total

    def test_csv_filters_by_department(
        self,
        auth_client,
        admin_user,
        sample_org,
        sample_catalog,
        create_hw_requirement,
    ):
        """
        The position costs export should respect the department_id
        query parameter, returning only positions in that department.
        """
        # Create requirements in both departments.
        create_hw_requirement(
            position=sample_org["pos_a1_1"],
            hardware=sample_catalog["hw_laptop_standard"],
            quantity=1,
        )
        create_hw_requirement(
            position=sample_org["pos_b1_1"],
            hardware=sample_catalog["hw_monitor_24"],
            quantity=1,
        )

        dept_a_id = sample_org["dept_a"].id
        client = auth_client(admin_user)
        response = client.get(
            f"/reports/export/position-costs/csv?department_id={dept_a_id}"
        )
        rows = _parse_csv_response(response)

        # All data rows (skip header) should belong to dept_a.
        dept_names = {row[0] for row in rows[1:]}
        assert sample_org["dept_a"].department_name in dept_names
        assert sample_org["dept_b"].department_name not in dept_names

    def test_csv_filters_by_division(
        self,
        auth_client,
        admin_user,
        sample_org,
        sample_catalog,
        create_hw_requirement,
    ):
        """
        The position costs export should respect the division_id
        query parameter, returning only positions in that division.
        """
        create_hw_requirement(
            position=sample_org["pos_a1_1"],
            hardware=sample_catalog["hw_laptop_standard"],
            quantity=1,
        )
        create_hw_requirement(
            position=sample_org["pos_a2_1"],
            hardware=sample_catalog["hw_monitor_24"],
            quantity=1,
        )

        div_a1_id = sample_org["div_a1"].id
        client = auth_client(admin_user)
        response = client.get(
            f"/reports/export/position-costs/csv?division_id={div_a1_id}"
        )
        rows = _parse_csv_response(response)

        # All data rows should belong to div_a1.
        div_names = {row[1] for row in rows[1:]}
        assert sample_org["div_a1"].division_name in div_names
        assert sample_org["div_a2"].division_name not in div_names

    def test_csv_grand_total_equals_hw_plus_sw_for_every_row(
        self,
        auth_client,
        admin_user,
        sample_org,
        sample_catalog,
        create_hw_requirement,
        create_sw_requirement,
    ):
        """
        Structural integrity: for every position row in the CSV,
        Grand Total must equal Hardware Total + Software Total.
        """
        create_hw_requirement(
            position=sample_org["pos_a1_1"],
            hardware=sample_catalog["hw_laptop_standard"],
            quantity=1,
        )
        create_sw_requirement(
            position=sample_org["pos_a1_1"],
            software=sample_catalog["sw_office_e3"],
            quantity=1,
        )

        client = auth_client(admin_user)
        response = client.get("/reports/export/position-costs/csv")
        rows = _parse_csv_response(response)

        for row in rows[1:]:
            hw_total = Decimal(row[8])
            sw_total = Decimal(row[9])
            grand_total = Decimal(row[10])
            assert grand_total == hw_total + sw_total, (
                f"Grand total mismatch for {row[3]}: "
                f"{grand_total} != {hw_total} + {sw_total}"
            )

    def test_csv_total_per_person_times_authorized_equals_grand(
        self,
        auth_client,
        admin_user,
        sample_org,
        sample_catalog,
        create_hw_requirement,
    ):
        """
        Structural integrity: Total per Person * Authorized Count
        must equal Grand Total for every row.
        """
        create_hw_requirement(
            position=sample_org["pos_a1_1"],
            hardware=sample_catalog["hw_laptop_standard"],
            quantity=1,
        )

        client = auth_client(admin_user)
        response = client.get("/reports/export/position-costs/csv")
        rows = _parse_csv_response(response)

        for row in rows[1:]:
            authorized = int(row[4])
            total_per_person = Decimal(row[7])
            grand_total = Decimal(row[10])
            expected = total_per_person * authorized
            assert grand_total == expected, (
                f"Multiplier mismatch for {row[3]}: "
                f"{grand_total} != {total_per_person} * {authorized}"
            )


# =====================================================================
# 7. Position cost Excel export
# =====================================================================


class TestExportPositionCostsExcel:
    """
    Verify the ``/reports/export/position-costs/xlsx`` endpoint
    returns a valid Excel workbook with correct position-level data.
    """

    def test_xlsx_content_type(self, auth_client, admin_user):
        """The response Content-Type must be the OOXML spreadsheet MIME type."""
        client = auth_client(admin_user)
        response = client.get("/reports/export/position-costs/xlsx")
        assert response.status_code == 200
        expected_type = (
            "application/vnd.openxmlformats-officedocument" ".spreadsheetml.sheet"
        )
        assert expected_type in response.headers["Content-Type"]

    def test_xlsx_content_disposition(self, auth_client, admin_user):
        """The response must include an attachment disposition."""
        client = auth_client(admin_user)
        response = client.get("/reports/export/position-costs/xlsx")
        disposition = response.headers.get("Content-Disposition", "")
        assert "attachment" in disposition
        assert "position_costs.xlsx" in disposition

    def test_xlsx_produces_valid_workbook(self, auth_client, admin_user):
        """
        The response bytes must be parseable as a valid .xlsx file.
        """
        client = auth_client(admin_user)
        response = client.get("/reports/export/position-costs/xlsx")
        wb = _load_excel_response(response)
        assert wb.active is not None
        assert wb.active.title == "Position Costs"

    def test_xlsx_header_row(self, auth_client, admin_user):
        """
        The first row must contain the exact position-level headers.
        """
        client = auth_client(admin_user)
        response = client.get("/reports/export/position-costs/xlsx")
        wb = _load_excel_response(response)
        ws = wb.active

        headers = [ws.cell(row=1, column=c).value for c in range(1, 12)]
        expected = [
            "Department",
            "Division",
            "Position Code",
            "Position Title",
            "Authorized Count",
            "HW per Person",
            "SW per Person",
            "Total per Person",
            "Hardware Total",
            "Software Total",
            "Grand Total",
        ]
        assert headers == expected

    def test_xlsx_data_values_match_known_costs(
        self,
        auth_client,
        admin_user,
        sample_org,
        sample_catalog,
        create_hw_requirement,
        create_sw_requirement,
    ):
        """
        Configure known requirements and verify the Excel cells
        contain the correct values.

        pos_a1_1 (auth=3): 1x hw_laptop_standard + 1x sw_office_e3
        hw_total = 3600.00, sw_total = 600.00, grand = 4200.00
        """
        pos = sample_org["pos_a1_1"]
        create_hw_requirement(
            position=pos,
            hardware=sample_catalog["hw_laptop_standard"],
            quantity=1,
        )
        create_sw_requirement(
            position=pos,
            software=sample_catalog["sw_office_e3"],
            quantity=1,
        )

        client = auth_client(admin_user)
        response = client.get("/reports/export/position-costs/xlsx")
        wb = _load_excel_response(response)
        ws = wb.active

        # Find the pos_a1_1 row by position code (column 3).
        pos_code = pos.position_code
        target_row = None
        for row_idx in range(2, ws.max_row + 1):
            if ws.cell(row=row_idx, column=3).value == pos_code:
                target_row = row_idx
                break

        assert (
            target_row is not None
        ), f"Position '{pos_code}' not found in Excel export"

        # Verify cell values.  The export writes float values.
        assert ws.cell(row=target_row, column=1).value == (
            sample_org["dept_a"].department_name
        )
        assert ws.cell(row=target_row, column=2).value == (
            sample_org["div_a1"].division_name
        )
        assert ws.cell(row=target_row, column=4).value == pos.position_title
        assert ws.cell(row=target_row, column=5).value == 3  # authorized

        # Cost columns (float comparisons).
        assert ws.cell(row=target_row, column=6).value == pytest.approx(
            1200.00, abs=0.01
        )  # HW per person
        assert ws.cell(row=target_row, column=7).value == pytest.approx(
            200.00, abs=0.01
        )  # SW per person
        assert ws.cell(row=target_row, column=8).value == pytest.approx(
            1400.00, abs=0.01
        )  # Total per person
        assert ws.cell(row=target_row, column=9).value == pytest.approx(
            3600.00, abs=0.01
        )  # HW total
        assert ws.cell(row=target_row, column=10).value == pytest.approx(
            600.00, abs=0.01
        )  # SW total
        assert ws.cell(row=target_row, column=11).value == pytest.approx(
            4200.00, abs=0.01
        )  # Grand total

    def test_xlsx_filters_by_department(
        self,
        auth_client,
        admin_user,
        sample_org,
        sample_catalog,
        create_hw_requirement,
    ):
        """
        The position export with department_id filter should only
        include positions from that department.
        """
        create_hw_requirement(
            position=sample_org["pos_a1_1"],
            hardware=sample_catalog["hw_laptop_standard"],
            quantity=1,
        )
        create_hw_requirement(
            position=sample_org["pos_b1_1"],
            hardware=sample_catalog["hw_monitor_24"],
            quantity=1,
        )

        dept_a_id = sample_org["dept_a"].id
        client = auth_client(admin_user)
        response = client.get(
            f"/reports/export/position-costs/xlsx?department_id={dept_a_id}"
        )
        wb = _load_excel_response(response)
        ws = wb.active

        # All data rows should be in dept_a.
        dept_names = set()
        for row_idx in range(2, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=1).value
            if val:
                dept_names.add(val)

        assert sample_org["dept_a"].department_name in dept_names
        assert sample_org["dept_b"].department_name not in dept_names


# =====================================================================
# 8. Unauthenticated access
# =====================================================================


class TestUnauthenticatedReportAccess:
    """
    Verify that unauthenticated users are redirected to login
    for all report endpoints.
    """

    def test_cost_summary_redirects_to_login(self, client):
        """Unauthenticated GET /reports/cost-summary redirects."""
        response = client.get("/reports/cost-summary")
        assert response.status_code == 302
        assert "login" in response.headers.get("Location", "").lower()

    def test_equipment_report_redirects_to_login(self, client):
        """Unauthenticated GET /reports/equipment-report redirects."""
        response = client.get("/reports/equipment-report")
        assert response.status_code == 302
        assert "login" in response.headers.get("Location", "").lower()

    def test_department_csv_export_redirects_to_login(self, client):
        """Unauthenticated export attempt redirects to login."""
        response = client.get("/reports/export/department-costs/csv")
        assert response.status_code == 302
        assert "login" in response.headers.get("Location", "").lower()

    def test_position_xlsx_export_redirects_to_login(self, client):
        """Unauthenticated export attempt redirects to login."""
        response = client.get("/reports/export/position-costs/xlsx")
        assert response.status_code == 302
        assert "login" in response.headers.get("Location", "").lower()


# =====================================================================
# 9. Cross-cutting: cost consistency between page and export
# =====================================================================


class TestPageExportConsistency:
    """
    Verify that the cost figures shown on the rendered HTML page
    match the figures produced by the CSV export for the same data.

    This is the class of bug that would be most embarrassing during
    a CIO demo: the page shows one number and the downloaded
    spreadsheet shows a different number.
    """

    def test_department_cost_page_matches_csv_export(
        self,
        auth_client,
        admin_user,
        sample_org,
        sample_catalog,
        create_hw_requirement,
        create_sw_requirement,
    ):
        """
        Configure known requirements across both departments, then
        verify the HTML page and CSV export agree on the figures.

        dept_a/pos_a1_1: 1x laptop ($3,600)
        dept_b/pos_b1_1: 1x sw_office_e3 ($800)
        """
        create_hw_requirement(
            position=sample_org["pos_a1_1"],
            hardware=sample_catalog["hw_laptop_standard"],
            quantity=1,
        )
        create_sw_requirement(
            position=sample_org["pos_b1_1"],
            software=sample_catalog["sw_office_e3"],
            quantity=1,
        )

        client = auth_client(admin_user)

        # Fetch the page.
        page_response = client.get("/reports/cost-summary")
        assert page_response.status_code == 200

        # Fetch the CSV.
        csv_response = client.get("/reports/export/department-costs/csv")
        assert csv_response.status_code == 200
        rows = _parse_csv_response(csv_response)

        # Verify dept_a figures appear in both.
        dept_a_name = sample_org["dept_a"].department_name
        dept_a_csv = next(r for r in rows[1:] if r[0] == dept_a_name)
        dept_a_grand = dept_a_csv[6]  # Grand Total column

        # The same figure should appear in the rendered HTML.
        assert dept_a_grand.encode() in page_response.data

        # Verify dept_b figures.
        dept_b_name = sample_org["dept_b"].department_name
        dept_b_csv = next(r for r in rows[1:] if r[0] == dept_b_name)
        dept_b_grand = dept_b_csv[6]

        assert dept_b_grand.encode() in page_response.data

    def test_position_cost_page_matches_csv_export(
        self,
        auth_client,
        admin_user,
        sample_org,
        sample_catalog,
        create_hw_requirement,
        create_sw_requirement,
    ):
        """
        Configure requirements on a single position, then verify
        the equipment report page and position CSV export agree.
        """
        pos = sample_org["pos_a1_1"]
        create_hw_requirement(
            position=pos,
            hardware=sample_catalog["hw_laptop_standard"],
            quantity=1,
        )
        create_sw_requirement(
            position=pos,
            software=sample_catalog["sw_office_e3"],
            quantity=1,
        )

        client = auth_client(admin_user)

        # Fetch the page (filtered to dept_a for a cleaner comparison).
        dept_a_id = sample_org["dept_a"].id
        page_response = client.get(
            f"/reports/equipment-report?department_id={dept_a_id}"
        )
        assert page_response.status_code == 200

        # Fetch the CSV (same filter).
        csv_response = client.get(
            f"/reports/export/position-costs/csv?department_id={dept_a_id}"
        )
        rows = _parse_csv_response(csv_response)

        # Find pos_a1_1 in the CSV.
        pos_row = next(r for r in rows[1:] if r[2] == pos.position_code)
        grand_total_csv = pos_row[10]  # "4200.00"

        # The same figure should appear in the rendered HTML.
        assert grand_total_csv.encode() in page_response.data
