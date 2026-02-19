"""
Export service — generate CSV and Excel files from report data.

All export functions return a BytesIO buffer ready to be sent as
a Flask response with the appropriate content type.
"""

import csv
import io
import logging
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.services import cost_service

logger = logging.getLogger(__name__)

# Excel header styling constants.
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", wrap_text=True)
_CURRENCY_FORMAT = '#,##0.00'


# =========================================================================
# CSV Exports
# =========================================================================

def export_department_costs_csv(
    department_summaries: list[cost_service.DepartmentCostSummary],
) -> io.BytesIO:
    """
    Export department-level cost summaries to CSV.

    Args:
        department_summaries: List of DepartmentCostSummary dataclasses.

    Returns:
        BytesIO buffer containing the CSV data.
    """
    output = io.StringIO()
    writer = csv.writer(output)

    # Write header row.
    writer.writerow([
        "Department",
        "Divisions",
        "Positions",
        "Authorized Headcount",
        "Hardware Total",
        "Software Total",
        "Grand Total",
    ])

    # Write data rows.
    for dept in department_summaries:
        writer.writerow([
            dept.department_name,
            dept.division_count,
            dept.position_count,
            dept.total_authorized,
            _format_decimal(dept.hardware_total),
            _format_decimal(dept.software_total),
            _format_decimal(dept.grand_total),
        ])

    # Convert to bytes for Flask response.
    buffer = io.BytesIO()
    buffer.write(output.getvalue().encode("utf-8-sig"))
    buffer.seek(0)
    return buffer


def export_position_costs_csv(
    position_summaries: list[cost_service.PositionCostSummary],
) -> io.BytesIO:
    """
    Export position-level cost details to CSV.

    Args:
        position_summaries: List of PositionCostSummary dataclasses.

    Returns:
        BytesIO buffer containing the CSV data.
    """
    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow([
        "Department",
        "Division",
        "Position Code",
        "Position Title",
        "Authorized Count",
        "Hardware per Person",
        "Software per Person",
        "Total per Person",
        "Hardware Total",
        "Software Total",
        "Grand Total",
    ])

    for pos in position_summaries:
        writer.writerow([
            pos.department_name,
            pos.division_name,
            pos.position_code,
            pos.position_title,
            pos.authorized_count,
            _format_decimal(pos.hardware_total_per_person),
            _format_decimal(pos.software_total_per_person),
            _format_decimal(pos.total_per_person),
            _format_decimal(pos.hardware_total),
            _format_decimal(pos.software_total),
            _format_decimal(pos.grand_total),
        ])

    buffer = io.BytesIO()
    buffer.write(output.getvalue().encode("utf-8-sig"))
    buffer.seek(0)
    return buffer


# =========================================================================
# Excel Exports
# =========================================================================

def export_department_costs_excel(
    department_summaries: list[cost_service.DepartmentCostSummary],
) -> io.BytesIO:
    """
    Export department-level cost summaries to an Excel workbook.

    Args:
        department_summaries: List of DepartmentCostSummary dataclasses.

    Returns:
        BytesIO buffer containing the .xlsx data.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Department Costs"

    headers = [
        "Department",
        "Divisions",
        "Positions",
        "Authorized Headcount",
        "Hardware Total",
        "Software Total",
        "Grand Total",
    ]
    _write_header_row(ws, headers)

    for row_idx, dept in enumerate(department_summaries, start=2):
        ws.cell(row=row_idx, column=1, value=dept.department_name)
        ws.cell(row=row_idx, column=2, value=dept.division_count)
        ws.cell(row=row_idx, column=3, value=dept.position_count)
        ws.cell(row=row_idx, column=4, value=dept.total_authorized)
        ws.cell(row=row_idx, column=5, value=float(dept.hardware_total)).number_format = _CURRENCY_FORMAT
        ws.cell(row=row_idx, column=6, value=float(dept.software_total)).number_format = _CURRENCY_FORMAT
        ws.cell(row=row_idx, column=7, value=float(dept.grand_total)).number_format = _CURRENCY_FORMAT

    _auto_fit_columns(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_position_costs_excel(
    position_summaries: list[cost_service.PositionCostSummary],
) -> io.BytesIO:
    """
    Export position-level cost details to an Excel workbook.

    Args:
        position_summaries: List of PositionCostSummary dataclasses.

    Returns:
        BytesIO buffer containing the .xlsx data.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Position Costs"

    headers = [
        "Department",
        "Division",
        "Position Code",
        "Position Title",
        "Authorized Count",
        "HW per Person",
        "SW per Person",
        "Total per Person",
        "Hardware Total",
        "Software Total",
        "Grand Total",
    ]
    _write_header_row(ws, headers)

    for row_idx, pos in enumerate(position_summaries, start=2):
        ws.cell(row=row_idx, column=1, value=pos.department_name)
        ws.cell(row=row_idx, column=2, value=pos.division_name)
        ws.cell(row=row_idx, column=3, value=pos.position_code)
        ws.cell(row=row_idx, column=4, value=pos.position_title)
        ws.cell(row=row_idx, column=5, value=pos.authorized_count)
        ws.cell(row=row_idx, column=6, value=float(pos.hardware_total_per_person)).number_format = _CURRENCY_FORMAT
        ws.cell(row=row_idx, column=7, value=float(pos.software_total_per_person)).number_format = _CURRENCY_FORMAT
        ws.cell(row=row_idx, column=8, value=float(pos.total_per_person)).number_format = _CURRENCY_FORMAT
        ws.cell(row=row_idx, column=9, value=float(pos.hardware_total)).number_format = _CURRENCY_FORMAT
        ws.cell(row=row_idx, column=10, value=float(pos.software_total)).number_format = _CURRENCY_FORMAT
        ws.cell(row=row_idx, column=11, value=float(pos.grand_total)).number_format = _CURRENCY_FORMAT

    _auto_fit_columns(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# =========================================================================
# Internal helpers
# =========================================================================

def _write_header_row(ws, headers: list[str]) -> None:
    """Write a styled header row to an Excel worksheet."""
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN


def _auto_fit_columns(ws) -> None:
    """Auto-fit column widths based on content (approximate)."""
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 4, 40)


def _format_decimal(value: Decimal) -> str:
    """Format a Decimal for CSV output."""
    return f"{value:.2f}"
